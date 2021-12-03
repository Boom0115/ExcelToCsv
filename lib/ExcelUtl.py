import pathlib
import openpyxl
from pprint import pprint


class ExcelUtl:

    def __init__(self):
        """
        クラスの初期化の時に呼ばれる処理
        """
        print("ExcelUtl is start")
        pass

    def ReadExcel(self, filename):
        # ファイルの存在チェック
        if pathlib.Path(filename).is_file() is False:
            raise FileNotFoundError(f"ファイルは存在していない  {filename}")

        workbook = openpyxl.load_workbook(filename)
        print(f"Excel file is Open :{filename}")
        for sheet_name in workbook.sheetnames:
            print(f"Sheet : {sheet_name}")

        sheet = workbook['Sheet1']

        header_row = 2

        start_row = 3
        start_col = 2
        end_col = 9

        # ヘッダ行の取得
        header = []
        for col in range(start_col, end_col + 1):
            header.append(str(sheet.cell(header_row, col).value))

        rows = list()

        for row in range(start_row, sheet.max_row + 1):
            # print(f"いま{row}行目")
            data = []
            for col in range(start_col, end_col + 1):
                # print(f"  いま{col}列目")
                val = sheet.cell(row, col).value
                if col == 2:
                    if type(val) != 'int':
                        raise ValueError(f"IDは数字でなけれならな:{val}")
                data.append(str(val))
            print(data)
            rows.append(data)

        # pprint(rows)

        print("Excel読み込み終了")

        o_file = r'C:\pythonProject\ExcelToCsv\TestExcel.csv'

        with open(o_file, 'w', encoding='utf8') as csv_f:
            csv_f.write('"' + '","'.join(header) + '"\n')
            for data in rows:
                csv = '"' + '","'.join(data) + '"\n'
                print(csv, end="")
                csv_f.write(csv)

if __name__ == '__main__':
    obj = ExcelUtl()
    excel_file = r'C:\pythonProject\ExcelToCsv\TestExcel.xlsx'
    obj.ReadExcel(excel_file)
